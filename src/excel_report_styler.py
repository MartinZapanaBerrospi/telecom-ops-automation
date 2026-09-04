# -*- coding: utf-8 -*-
"""
Generador y Estilizador Corporativo de Reportes Excel para Jefatura
Autor: Martín Zapana Berrospi
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

def generar_reporte_ejecutivo_estilizado():
    excel_path = DATA_DIR / "Reporte_Gerencial_Facturacion_Estilizado.xlsx"
    print(f"[*] Generando reporte gerencial estilizado en: {excel_path}")

    # Datos de resumen
    data_kpi = {
        "Indicador Operacional": [
            "Facturación Total Emitida (S/)",
            "Recaudación Efectiva Cobrada (S/)",
            "Cartera en Mora / Vencida (S/)",
            "Notas de Crédito y Ajustes (S/)",
            "Tasa de Error / Disputas (%)",
            "Efectividad de Cobranza (%)",
            "Total Recibos Auditados"
        ],
        "Meta / Target": ["S/ 1,200,000", "S/ 1,000,000", "< S/ 150,000", "< S/ 25,000", "< 2.0 %", "> 85.0 %", "20,000"],
        "Resultado Actual": ["S/ 1,348,920", "S/ 1,146,582", "S/ 162,338", "S/ 18,450", "1.37 %", "85.00 %", "20,429"],
        "Estado": ["CUMPLIDO", "CUMPLIDO", "EN OBSERVACION", "CUMPLIDO", "CUMPLIDO", "CUMPLIDO", "CUMPLIDO"]
    }
    df_kpi = pd.DataFrame(data_kpi)

    # Detalle de ciclos
    data_ciclos = {
        "Ciclo": ["C01 (Corte 01)", "C15 (Corte 15)", "C28 (Corte 28)"],
        "Recibos Emitidos": [6810, 6805, 6814],
        "Monto Facturado (S/)": [448200.0, 451320.0, 449400.0],
        "Ajustes Aprobados (S/)": [5820.0, 6410.0, 6220.0],
        "Cobranza (%)": [86.2, 84.8, 85.1]
    }
    df_ciclos = pd.DataFrame(data_ciclos)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_kpi.to_excel(writer, sheet_name="Resumen_KPIs", index=False, startrow=3)
        df_ciclos.to_excel(writer, sheet_name="Detalle_Ciclos", index=False, startrow=3)

    # Aplicar estilos con openpyxl
    wb = openpyxl.load_workbook(excel_path)
    
    # Paleta Corporativa Telecom
    header_color = "005A9E"  # Telecom Enterprise Blue
    dark_navy = "0F172A"
    light_gray = "F1F5F9"
    green_ok = "DCFCE7"
    yellow_warn = "FEF3C7"

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True
        
        # Título Corporativo
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = f"TELECOM ENTERPRISE — REPORTE OPERACIONAL DE POST FACTURACIÓN ({sheetname.replace('_', ' ')})"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Subtítulo con fecha
        ws.merge_cells("A2:D2")
        sub_cell = ws["A2"]
        sub_cell.value = "Generado automáticamente por Pipeline de Automatización — Confidencial Operaciones"
        sub_cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # Encabezados de tabla (Fila 4)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=dark_navy, end_color=dark_navy, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 24

        # Filas de datos
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for row_idx in range(5, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                c.font = Font(name="Calibri", size=10)
                if col_idx == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

                # Pintar estados condicionalmente
                if c.value == "CUMPLIDO":
                    c.fill = PatternFill(start_color=green_ok, end_color=green_ok, fill_type="solid")
                    c.font = Font(name="Calibri", size=10, bold=True, color="166534")
                elif c.value == "EN OBSERVACION":
                    c.fill = PatternFill(start_color=yellow_warn, end_color=yellow_warn, fill_type="solid")
                    c.font = Font(name="Calibri", size=10, bold=True, color="92400E")

        # Auto-ajuste de ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(excel_path)
    print(f"[OK] Reporte con formato corporativo guardado exitosamente.")

if __name__ == "__main__":
    generar_reporte_ejecutivo_estilizado()
